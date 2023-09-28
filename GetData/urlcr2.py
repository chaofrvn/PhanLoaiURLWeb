import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os

def extract_urls(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    urls = set()
    for link in soup.find_all('a', href=True):
        href = link['href']
        if not href.startswith('http'):
            href = url + href
        urls.add(href)
    return urls



def save_to_excel(urls, output_file):
    if os.path.exists(output_file):  # Kiểm tra xem tệp Excel đã tồn tại hay chưa.
        workbook = load_workbook(output_file)  # Nếu đã tồn tại, mở tệp Excel lên.
        if "URLs" in workbook.sheetnames:  # Kiểm tra xem trang tính "URLs" đã tồn tại trong tệp Excel hay chưa.
            sheet = workbook["URLs"]  # Nếu đã tồn tại, mở trang tính "URLs" lên.
        else:
            sheet = workbook.active  # Nếu chưa tồn tại, tạo một trang tính mới.
            sheet.title = "URLs"
            sheet.append(["URL"])  # Thêm tiêu đề "URL" vào ô đầu tiên của trang tính mới.
    else:
        workbook = Workbook()  # Nếu tệp Excel chưa tồn tại, tạo một tệp Excel mới.
        sheet = workbook.active
        sheet.title = "URLs"
        sheet.append(["URL"])

    for url in urls:
        sheet.append([url])

    workbook.save(output_file)

if __name__ == "__main__":
    target_url = input("Enter the URL to scrape: ")
    output_file = input("Enter the output Excel file name (e.g., output.xlsx): ")
    # output_file = "phapluat.xlsx"

    urls = extract_urls(target_url)
    save_to_excel(urls, output_file)

    print(f"URLs saved to {output_file}")